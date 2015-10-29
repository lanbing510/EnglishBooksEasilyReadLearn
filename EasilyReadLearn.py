# -*- coding: utf-8 -*-
"""
Created on Sun Oct 25 09:12:17 2015
Function: Generate Word-Familiarity-Frequence-Explain Tables
@Author: lanbing510
"""

import os
import re
import sys
import sqlite3
import openpyxl
reload(sys)
sys.setdefaultencoding('utf8')

os.chdir(sys.path[0])

db=sqlite3.connect("english_dictionary_sqlite.db")
db.text_factory=str
cu=db.cursor()

def generate_word_frequence(filename):
    file=open(filename,"r")
    strstream=""
    for line in file.xreadlines():
        strstream+=line
    strstream=re.sub(r"(-\n|\:)","",strstream)
    strstream=re.sub(r"(\n{1,}|,|\.|\(|\)|\[|\]|\{|\}|\')"," ",strstream)
    word_list=re.findall(r"[A-Za-z]{2,}",strstream)    
    word_frequence={}
    for w in word_list:
        if word_frequence.has_key(w):
            word_frequence[w]+=1;
        else:
            word_frequence.update({w.lower():1})
    return word_frequence

def lookup_sql(word):
    command="select explain from english_dictionary where word='%s'" % word
    cu.execute(command)
    return cu.fetchone()

def lookup_dictionary(word):
    possible_words=[]
    possible_words.append(word)
    word_length=len(word)
    if word[-1]=="s":
        possible_words.append(word[0:-1])
    elif word_length>1 and word[-2:]=="ed":
        possible_words.append(word[0:-2])
        possible_words.append(word[0:-1])
    elif word_length>2 and word[-3:]=="ing":
        possible_words.append(word[0:-3])
        possible_words.append(word[0:-3]+"e")
    else:
        pass
    got=False
    explain=None
    for pw in possible_words:
        explain=lookup_sql(pw)
        if explain:
            got=True
            word=pw
            break  
    if got:
        return (word,explain[0])
    else:
        return (word,None)    

def generate_word_table(word_frequence,sheet_title):
    wb=openpyxl.Workbook(optimized_write=True)
    excel_output_filename="tolearn.xlsx"
    if not os.path.isfile(excel_output_filename):
        wb.save(excel_output_filename)
    wb=openpyxl.load_workbook(filename=excel_output_filename)
    ws=wb.create_sheet(title=sheet_title)
    ws.append([u"单词",u"生疏度",u"频率",u"解释"])
    row_count=1;
    for wd in word_frequence:
        (word,explain)=lookup_dictionary(wd)
        if explain:
            row_count+=1
            ws.append([word,1,word_frequence.get(wd)])
            ws.cell(row=row_count,column=4).set_explicit_value(value=explain,data_type="s")
            #ws.append([word,1,word_frequence.get(wd),explain])
    if "Sheet1" in wb.get_sheet_names():
        wb.remove_sheet(wb["Sheet1"])
    wb.save(excel_output_filename)


if __name__=="__main__":
    if len(sys.argv)<2:
        print u"Usage: EasilyReadLearn.py {Your PDF File} [Start Page Number-Stop Page Number]"
        exit()
    file_prefix=sys.argv[1].split("\\")[-1][:-4]
    if len(file_prefix)>31:
        print u"文件名字太长，请缩减"
        exit()
    if not os.path.isfile(sys.argv[1]):
        print u"没有找到%s.pdf这个文件，请确定路径" % file_prefix
        exit()
    print u"开始处理..."
    syscmd=""
    if len(sys.argv)==2:
        syscmd="pdf2txt -o %s.txt %s " % (sys.path[0]+"\\"+file_prefix,sys.argv[1])
    elif len(sys.argv)==3:
        file_prefix+=sys.argv[2]
        syscmd="pdf2txt -o %s.txt -p " % file_prefix
        page_range=sys.argv[2].split("-")
        for i in range(int(page_range[0]),int(page_range[1])):
            syscmd+=str(i)+","
        syscmd=syscmd[:-1]+" "
        syscmd+=sys.argv[1]
    os.system(syscmd)
    word_frequence=generate_word_frequence(file_prefix+".txt")
    generate_word_table(word_frequence,file_prefix)
    print u"已经成功将单词表输入到了tolearn.xlsx :)"
    print u"需要去除最常用单词吗？是：Y,否：N"
    flag=raw_input()
    if flag.lower()=="y":
        os.system("python -W ignore Sync.py")
    print u"完成，可以打开tolearn.xlsx进行学习了了:)"
    syscmd="explorer %s" % sys.path[0]
    os.system(syscmd)