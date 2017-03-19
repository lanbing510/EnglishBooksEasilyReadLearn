# -*- coding: utf-8 -*-
"""
Created on Wed Oct 28 21:46:19 2015
Function: Sync Data Between Tolearn.xlsl and Mastered.xlsl
@Author: lanbing510
"""

import os
import openpyxl
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

os.chdir(sys.path[0])

wb_tolearn=openpyxl.load_workbook(filename="tolearn.xlsm",keep_vba=True)
wb_mastered=openpyxl.load_workbook(filename="mastered.xlsm",keep_vba=True)

dict_mastered={}
ws_mastered=wb_mastered['mastered']
row_count=0
for row in ws_mastered.rows:
    row_count+=1
    if row_count>1:
        dict_mastered.update({row[0].value:row_count})


# tolearn.xslx -> mastered.xslx
for sheet in wb_tolearn.get_sheet_names():
    ws_tolearn=wb_tolearn[sheet]
    row_count=0
    for row in ws_tolearn.rows:
        row_count+=1
        if row_count>1:
            if row[0].value in dict_mastered:
                ws_mastered.cell(row=dict_mastered[row[0].value],column=3).value+=row[2].value
                for i in range(9):
                    ws_tolearn.cell(row=row_count,column=i+1).value=None
            elif row[1].value==0:
                ws_mastered.append([row[0].value,0,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value])
                dict_mastered.update({row[0].value:ws_mastered.get_highest_row()})
                for i in range(9):
                    ws_tolearn.cell(row=row_count,column=i+1).value=None
            else:
                pass

# mastered.xslx -> tolearn.xslx
for sheet in wb_mastered.get_sheet_names():
    ws_mastered=wb_mastered[sheet]
    row_count=0
    for row in ws_mastered.rows:
        row_count+=1
        if row_count>1:
            if row[1].value>0:
                if "from_mastered" not in wb_tolearn.get_sheet_names():
                    wb_tolearn.create_sheet(title="from_mastered")
                    wb_tolearn["from_mastered"].append([u"单词",u"生疏度",u"频率",u"解释",u"英式音标",u"美式音标",u"英式发音",u"美式发音",u"例句"])
                wb_tolearn["from_mastered"].append([row[0].value,row[1].value,row[2].value])
                row_index=wb_tolearn["from_mastered"].get_highest_row()   
                for i in range(6):
                    wb_tolearn["from_mastered"].cell(row=row_index,column=4+i).set_explicit_value(value=row[3+i].value,data_type="s")
                for i in range(9):
                    ws_mastered.cell(row=row_count,column=i+1).value=None
            elif row[1].value==-1:
                for i in range(9):
                    ws_mastered.cell(row=row_count,column=i+1).value=None
            else:
                pass


# remove the none areas to polish the sheets
def polish_sheets(wb):
    for sheet in wb.get_sheet_names():
        ws=wb[sheet]
        ws_filtered=wb.create_sheet(title=sheet+"_temp")
        row_count=0
        for row in ws.rows:
            if row[0].value and row[1].value!=-1:
                row_count+=1
                ws_filtered.append([row[0].value,row[1].value,row[2].value])
                for i in range(6):
                    ws_filtered.cell(row=row_count,column=4+i).set_explicit_value(value=row[3+i].value,data_type="s")
        wb.remove_sheet(wb[sheet])
        wb[sheet+"_temp"].title=sheet
        if row_count==1: # remove the empty sheet
            wb.remove_sheet(wb[sheet])


#polish_sheets(wb_tolearn)
#polish_sheets(wb_mastered)

wb_tolearn.save("tolearn.xlsm")
wb_mastered.save("mastered.xlsm")